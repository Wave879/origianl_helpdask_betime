#!/usr/bin/env python3
"""
Import the employee workbook into the BETIME users table.

The workbook is treated as the source of truth for the employee directory:
- email2 is used as the primary login/email when available;
- a stable fallback email is generated for rows that do not have one;
- new users receive a temporary password and must change it on first login.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2


DEFAULT_EXCEL = Path(r"C:\Users\wave\Downloads\ข้อมูลพนักงานเพื่อนำไปทำ ToolหรับAi backoffice.xlsx")
DEFAULT_PG_URL = os.environ.get("PG_URL", "postgres://postgres:123456@localhost:5432/Betime_DB?client_encoding=utf8")
EMAIL_FALLBACK_DOMAIN = "betimes.local"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\r", "\n").replace("\t", " ").strip()
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def slugify(value: str) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9ก-๙]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "employee"


def generate_temp_password(length: int = 10) -> str:
    alphabet = "ABCDEFGHJKMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz23456789"
    return "".join(alphabet[b % len(alphabet)] for b in os.urandom(length))


def sha256_hex(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def parse_tags(value: Any) -> list[str]:
    if isinstance(value, list):
        seen = []
        for item in value:
          text = clean_text(item)
          if text and text not in seen:
              seen.append(text)
        return seen
    text = clean_text(value)
    if not text:
        return []
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            seen = []
            for item in parsed:
                text_item = clean_text(item)
                if text_item and text_item not in seen:
                    seen.append(text_item)
            return seen
    except Exception:
        pass
    seen = []
    for item in text.split(","):
        text_item = clean_text(item)
        if text_item and text_item not in seen:
            seen.append(text_item)
    return seen


def serialize_tags(value: Any) -> str:
    return json.dumps(parse_tags(value), ensure_ascii=False)


def add_tag(value: Any, tag: str) -> str:
    tags = parse_tags(value)
    normalized = clean_text(tag)
    if normalized and normalized not in tags:
        tags.append(normalized)
    return json.dumps(tags, ensure_ascii=False)


def pick_email(row_number: int, full_name: str, email2: str, email1: str) -> str:
    for candidate in (email2, email1):
        text = clean_text(candidate).strip().lower()
        if "@" in text:
            return text
    base = slugify(full_name or f"employee-{row_number}")
    return f"{base}-{row_number}@{EMAIL_FALLBACK_DOMAIN}"


def pick_role(level: str, title: str) -> str:
    level_l = clean_text(level).lower()
    title_l = clean_text(title).lower()
    if "ceo" in level_l or "chief executive" in title_l:
        return "ceo"
    if any(token in level_l for token in ("director", "manager", "assistant manager")):
        return "manager"
    if any(token in title_l for token in ("director", "manager", "chief")):
        return "manager"
    return "staff"


def ensure_schema(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
              id TEXT PRIMARY KEY,
              email TEXT UNIQUE NOT NULL,
              username TEXT UNIQUE,
              full_name TEXT,
              password_hash TEXT,
              role TEXT DEFAULT 'staff',
              department TEXT,
              avatar_url TEXT,
              is_active INTEGER DEFAULT 1,
              must_change_password INTEGER DEFAULT 0,
              tags TEXT DEFAULT '[]',
              access_mode TEXT DEFAULT 'role',
              access_json TEXT DEFAULT '[]',
              created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
              updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS tags TEXT DEFAULT '[]'")
    conn.commit()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="Path to the employee workbook")
    parser.add_argument("--pg-url", default=DEFAULT_PG_URL, help="PostgreSQL connection string")
    parser.add_argument("--dry-run", action="store_true", help="Parse workbook but do not write to the database")
    parser.add_argument(
        "--reset-existing-passwords",
        action="store_true",
        help="Regenerate temporary passwords for existing users too",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    skipped_blank = 0
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row) + [None] * (18 - len(row))
        full_name = clean_text(values[5])
        if not full_name and not clean_text(values[17]) and not clean_text(values[16]):
            skipped_blank += 1
            continue

        nickname = clean_text(values[0])
        company = clean_text(values[1])
        position = clean_text(values[11])
        level = clean_text(values[12])
        department = clean_text(values[13]) or clean_text(values[14])
        manager = clean_text(values[15])
        email_display = clean_text(values[16])
        email = pick_email(row_number, full_name or nickname, clean_text(values[17]), email_display)
        role = pick_role(level, position)
        temp_password = generate_temp_password()

        records.append(
            {
                "row_number": row_number,
                "email": email,
                "username": email,
                "full_name": full_name or nickname or email,
                "department": department,
                "role": role,
                "position": position,
                "level": level,
                "manager": manager,
                "company": company,
                "nickname": nickname,
                "temp_password": temp_password,
            }
        )

    print(f"Parsed {len(records)} employee rows from {excel_path.name}")
    print(f"Skipped blank rows: {skipped_blank}")
    if records:
        print(f"First employee: {records[0]['full_name']} <{records[0]['email']}>")

    if args.dry_run:
        return 0

    conn = psycopg2.connect(args.pg_url)
    try:
        ensure_schema(conn)
        inserted = 0
        updated = 0
        reset_pw = 0
        report_rows = []
        with conn.cursor() as cur:
            for record in records:
                password_hash = sha256_hex(record["temp_password"])
                cur.execute(
                    """
                    SELECT id, password_hash, tags
                    FROM users
                    WHERE LOWER(email) = LOWER(%s)
                    LIMIT 1
                    """,
                    [record["email"]],
                )
                existing = cur.fetchone()
                if existing:
                    updated += 1
                    tags_json = add_tag(existing[2] if len(existing) > 2 else None, "รอรับรหัส ครั้งแรก")
                    if args.reset_existing_passwords:
                        cur.execute(
                            """
                            UPDATE users
                            SET username=%s,
                                full_name=%s,
                                role=%s,
                                department=%s,
                                password_hash=%s,
                                must_change_password=1,
                                tags=%s,
                                is_active=1,
                                updated_at=now()
                            WHERE id=%s
                            """,
                            [
                                record["username"],
                                record["full_name"],
                                record["role"],
                                record["department"],
                                password_hash,
                                tags_json,
                                existing[0],
                            ],
                        )
                        reset_pw += 1
                        report_rows.append(record)
                    else:
                        cur.execute(
                            """
                            UPDATE users
                            SET username=%s,
                                full_name=%s,
                                role=%s,
                                department=%s,
                                tags=%s,
                                is_active=1,
                                updated_at=now()
                            WHERE id=%s
                            """,
                            [
                                record["username"],
                                record["full_name"],
                                record["role"],
                                record["department"],
                                tags_json,
                                existing[0],
                            ],
                        )
                    continue

                cur.execute(
                    """
                    INSERT INTO users
                      (id, email, username, full_name, password_hash, role, department, is_active, must_change_password, tags, access_mode, access_json)
                    VALUES
                      (%s, %s, %s, %s, %s, %s, %s, 1, 1, %s, %s, %s)
                    """,
                    [
                        f"usr_{record['row_number']}_{re.sub(r'[^a-z0-9]+', '', record['email'].lower())[:16] or 'employee'}",
                        record["email"],
                        record["username"],
                        record["full_name"],
                        password_hash,
                        record["role"],
                        record["department"],
                        serialize_tags(["รอรับรหัส ครั้งแรก"]),
                        'role',
                        '[]',
                    ],
                )
                inserted += 1
                report_rows.append(record)

        conn.commit()
        print(f"Inserted: {inserted}")
        print(f"Updated: {updated}")
        print(f"Password resets: {reset_pw}")
        print("Temporary password is generated per newly inserted user.")
        if report_rows:
            print("Credentials report:")
            for record in report_rows:
                print(f"- {record['full_name']} <{record['email']}> => {record['temp_password']}")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
