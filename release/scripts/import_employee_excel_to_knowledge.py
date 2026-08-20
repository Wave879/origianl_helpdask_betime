#!/usr/bin/env python3
"""
Import an employee directory Excel workbook into the Knowledge Center.

Each row becomes one knowledge article so the AI backoffice can search by
name, nickname, position, department, manager, and email.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path
from typing import Iterable, List

import openpyxl
import psycopg2
from psycopg2.extras import execute_batch


DEFAULT_EXCEL = Path(r"C:\Users\wave\Downloads\ข้อมูลพนักงานเพื่อนำไปทำ ToolหรับAi backoffice.xlsx")
DEFAULT_PG_URL = "postgres://postgres:123456@localhost:5432/Betime_DB"
DEFAULT_CATEGORY = "Employee Knowledge"
DEFAULT_AUTHOR = "Employee Excel Import"


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\r", "\n").replace("\t", " ").strip()
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def slugify(value: str) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9ก-๙]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "row"


def uniq(values: Iterable[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


def structured_tag(key: str, value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    return f"{key}:{text}"


def build_content(row: dict, row_number: int) -> tuple[str, str, str]:
    company = clean_text(row.get("บริษัท"))
    prefix = clean_text(row.get("คำนำหน้า"))
    first_name = clean_text(row.get("ชื่อ"))
    last_name = clean_text(row.get("นามสกุล"))
    full_name = clean_text(row.get("ชื่อเต็ม")) or " ".join(part for part in [prefix, first_name, last_name] if part).strip()
    nickname = clean_text(row.get("ชื่อเล่น"))
    first_en = clean_text(row.get("ชื่อ (EN)"))
    last_en = clean_text(row.get("นามสกุล (EN)"))
    nick_en = clean_text(row.get("ชื่อเล่น (EN)"))
    gender = clean_text(row.get("เพศ"))
    birthdate = clean_text(row.get("วันเกิด"))
    position = clean_text(row.get("ชื่อตำแหน่ง"))
    level = clean_text(row.get("ระดับ"))
    department = clean_text(row.get("ฝ่าย"))
    affiliation = clean_text(row.get("สังกัด"))
    manager = clean_text(row.get("ชื่อหัวหน้าภาษาไทย"))
    email1 = clean_text(row.get("อีเมลบริษัท"))
    email2 = clean_text(row.get("อีเมลบริษัท2"))

    title = full_name or nickname or email1 or f"Employee {row_number}"
    if position:
      title = f"{title} — {position}"

    content_lines = [
        f"ชื่อเต็ม: {full_name or '-'}",
        f"ชื่อเล่น: {nickname or '-'}",
        f"ชื่ออังกฤษ: {' '.join(part for part in [first_en, last_en] if part).strip() or '-'}",
        f"ชื่อเล่นอังกฤษ: {nick_en or '-'}",
        f"ตำแหน่ง: {position or '-'}",
        f"ระดับ: {level or '-'}",
        f"ฝ่าย: {department or '-'}",
        f"สังกัด: {affiliation or '-'}",
        f"หัวหน้า: {manager or '-'}",
        f"เพศ: {gender or '-'}",
        f"วันเกิด: {birthdate or '-'}",
        f"อีเมลบริษัท: {email1 or '-'}",
        f"อีเมลบริษัท2: {email2 or '-'}",
        f"บริษัท: {company or '-'}",
        f"แถวข้อมูลในไฟล์: {row_number}",
    ]

    tags = uniq([
        "employee",
        "backoffice",
        company,
        department,
        affiliation,
        level,
        position,
        manager,
        nickname,
        nick_en,
        email1,
        email2,
        gender,
        structured_tag("company", company),
        structured_tag("department", department),
        structured_tag("position", position),
        structured_tag("level", level),
        structured_tag("manager", manager),
        structured_tag("gender", gender),
    ])

    searchable_blobs = uniq([
        title,
        full_name,
        nickname,
        first_en,
        last_en,
        nick_en,
        position,
        level,
        department,
        affiliation,
        manager,
        email1,
        email2,
        company,
        gender,
        birthdate,
    ])

    return title, "\n".join(content_lines), ", ".join(tags + searchable_blobs[:6])


def ensure_schema(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS knowledge_articles (
              id TEXT PRIMARY KEY,
              title TEXT NOT NULL,
              content TEXT,
              category TEXT,
              tags TEXT DEFAULT '',
              author TEXT,
              created_at TIMESTAMPTZ DEFAULT now(),
              updated_at TIMESTAMPTZ DEFAULT now()
            )
            """
        )
    conn.commit()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="Path to the employee workbook")
    parser.add_argument("--pg-url", default=DEFAULT_PG_URL, help="PostgreSQL connection string")
    parser.add_argument("--category", default=DEFAULT_CATEGORY, help="Knowledge article category")
    parser.add_argument("--author", default=DEFAULT_AUTHOR, help="Article author name")
    parser.add_argument("--dry-run", action="store_true", help="Parse workbook but do not write to the database")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Workbook is empty")

    headers = [clean_text(value) for value in rows[0]]
    records = []
    for idx, values in enumerate(rows[1:], start=2):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        title, content, tags = build_content(row, idx)
        if not title.strip():
            continue
        records.append(
            {
                "id": f"employee_excel_{idx}_{slugify(title)}",
                "title": title,
                "content": content,
                "category": args.category,
                "tags": tags,
                "author": args.author,
            }
        )

    print(f"Parsed {len(records)} employee knowledge articles from {excel_path.name}")
    if records:
        print(f"First article: {records[0]['title']}")

    if args.dry_run:
        return 0

    conn = psycopg2.connect(args.pg_url)
    try:
        ensure_schema(conn)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM knowledge_articles WHERE id LIKE 'employee_excel_%'")
            execute_batch(
                cur,
                """
                INSERT INTO knowledge_articles
                (id, title, content, category, tags, author, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, now(), now())
                ON CONFLICT (id) DO UPDATE SET
                  title = EXCLUDED.title,
                  content = EXCLUDED.content,
                  category = EXCLUDED.category,
                  tags = EXCLUDED.tags,
                  author = EXCLUDED.author,
                  updated_at = now()
                """,
                [
                    (
                        row["id"],
                        row["title"],
                        row["content"],
                        row["category"],
                        row["tags"],
                        row["author"],
                    )
                    for row in records
                ],
                page_size=100,
            )
        conn.commit()
        print(f"Imported {len(records)} employee knowledge articles into PostgreSQL")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
